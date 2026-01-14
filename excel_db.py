import os
from openpyxl import Workbook, load_workbook
from app import config

HEADERS = ["Nombre", "Cédula", "Cargo", "Salario"]


def ensure_file():
    if not os.path.exists(config.DATA_EXCEL):
        wb = Workbook()
        ws = wb.active
        ws.title = "empleados"
        ws.append(HEADERS)
        wb.save(config.DATA_EXCEL)


def read_all():
    ensure_file()
    wb = load_workbook(config.DATA_EXCEL)
    ws = wb["empleados"]
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if i == 1:
            continue
        rows.append(list(row))
    return rows


def add_record(record):
    ensure_file()
    wb = load_workbook(config.DATA_EXCEL)
    ws = wb["empleados"]
    ws.append([record.get(h, "") for h in HEADERS])
    wb.save(config.DATA_EXCEL)


def delete_record(row_index):
    wb = load_workbook(config.DATA_EXCEL)
    ws = wb["empleados"]
    ws.delete_rows(row_index + 2)
    wb.save(config.DATA_EXCEL)


def update_record(row_index, record):
    wb = load_workbook(config.DATA_EXCEL)
    ws = wb["empleados"]
    r = row_index + 2
    for c, h in enumerate(HEADERS, start=1):
        ws.cell(row=r, column=c, value=record.get(h, ""))
    wb.save(config.DATA_EXCEL)
